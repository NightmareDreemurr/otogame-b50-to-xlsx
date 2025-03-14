import requests
import json
import time
import argparse
import os
import sys
from urllib.parse import urlparse, parse_qs
from bs4 import BeautifulSoup
import logging
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.cell import MergedCell
import cloudscraper
import traceback

# 设置日志
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("otogame_debug.log", encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def print_request_info(req):
    """打印请求信息，用于调试"""
    logger.debug("=== 请求信息 ===")
    logger.debug(f"请求URL: {req.url}")
    logger.debug(f"请求方法: {req.method}")
    logger.debug(f"请求头: {dict(req.headers)}")
    logger.debug(f"请求Cookies: {req._cookies.get_dict() if hasattr(req, '_cookies') else None}")
    if req.body:
        try:
            if isinstance(req.body, bytes):
                body = req.body.decode('utf-8')
            else:
                body = req.body
            logger.debug(f"请求体: {body[:500]}..." if len(str(body)) > 500 else f"请求体: {body}")
        except:
            logger.debug(f"请求体: [无法解码]")

def print_response_info(resp):
    """打印响应信息，用于调试"""
    logger.debug("=== 响应信息 ===")
    logger.debug(f"响应状态码: {resp.status_code}")
    logger.debug(f"响应头: {dict(resp.headers)}")
    logger.debug(f"响应Cookies: {requests.utils.dict_from_cookiejar(resp.cookies)}")
    
    # 尝试解析响应内容
    try:
        if resp.headers.get('Content-Type', '').startswith('application/json'):
            response_json = resp.json()
            logger.debug(f"响应体(JSON): {json.dumps(response_json, ensure_ascii=False, indent=2)}")
            # 如果是错误响应，特别记录错误信息
            if resp.status_code >= 400:
                logger.error(f"错误代码: {response_json.get('code')}")
                logger.error(f"错误消息: {response_json.get('message')}")
                logger.error(f"时间戳: {response_json.get('timestamp')}")
        else:
            logger.debug(f"响应体: {resp.text}")
    except Exception as e:
        logger.debug(f"响应体(原始): {resp.text}")
        logger.debug(f"解析响应体失败: {str(e)}")
    logger.debug("=== 响应信息结束 ===")

def create_session():
    """创建无代理会话"""
    # 使用cloudscraper创建会话
    session = cloudscraper.create_scraper(
        browser={
            'browser': 'chrome',
            'platform': 'windows',
            'desktop': True
        }
    )
    
    # 禁用代理设置
    session.proxies = {
        'http': None,
        'https': None
    }
    return session

def login_and_get_token(email, password):
    """
    完整的登录流程，从获取重定向URL到获取授权令牌
    """
    logger.info("启动登录流程...")
    session = create_session()
    
    # 设置用户代理
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36"
    
    try:
        # 第一步：从u.otogame.net获取重定向URL
        logger.info("Step 1: 获取OAuth重定向URL")
        redirect_api_url = "https://u.otogame.net/api/aime/user/redirect"
        headers = {
            "User-Agent": user_agent,
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Referer": "https://u.otogame.net/",
            "X-Requested-With": "XMLHttpRequest"
        }
        
        logger.debug(f"发送请求到: {redirect_api_url}")
        response = session.get(redirect_api_url, headers=headers, timeout=30)
        print_response_info(response)
        response.raise_for_status()
        
        try:
            redirect_data = response.json()
            if redirect_data.get("code") != 0:
                logger.error(f"获取重定向URL失败: {redirect_data.get('message', '未知错误')}")
                return None
                
            oauth_url = redirect_data.get("data", {}).get("redirect")
            if not oauth_url:
                logger.error("无法获取OAuth授权URL")
                return None
                
            logger.info(f"成功获取OAuth URL: {oauth_url}")
            
            # 第二步：访问OAuth URL，这会重定向到登录页面
            logger.info("Step 2: 访问OAuth授权页面")
            oauth_headers = {
                "User-Agent": user_agent,
                "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                "Referer": "https://u.otogame.net/",
                "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": "\"Windows\"",
                "sec-fetch-dest": "document",
                "sec-fetch-mode": "navigate",
                "sec-fetch-site": "cross-site",
                "sec-fetch-user": "?1",
                "upgrade-insecure-requests": "1"
            }
            
            logger.debug(f"发送请求到OAuth地址: {oauth_url}")
            oauth_response = session.get(oauth_url, headers=oauth_headers, allow_redirects=True, timeout=30)
            print_response_info(oauth_response)
            logger.debug(f"当前Cookies: {requests.utils.dict_from_cookiejar(session.cookies)}")
            
            # 检查OAuth重定向响应是否是授权确认页面
            if "授权提示" in oauth_response.text and "想要访问您的账户" in oauth_response.text:
                logger.info("检测到授权确认页面，正在处理...")
                
                # 解析HTML获取表单数据
                soup = BeautifulSoup(oauth_response.text, 'html.parser')
                # 查找同意按钮所在的表单（不含DELETE方法的表单）
                authorize_form = soup.find('form', {'action': lambda x: x and '/oauth/authorize' in x and not 'DELETE' in str(x)})
                
                if not authorize_form:
                    logger.error("无法找到授权表单")
                    return None
                    
                # 提取表单中的所有隐藏字段
                form_data = {}
                for input_tag in authorize_form.find_all('input', {'type': 'hidden'}):
                    if input_tag.get('name') and input_tag.get('value'):
                        form_data[input_tag['name']] = input_tag['value']
                
                logger.debug(f"授权表单数据: {form_data}")
                
                # 发送同意授权的请求
                authorize_url = "https://bemanicn.com/oauth/authorize"
                authorize_headers = {
                    "User-Agent": user_agent,
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Content-Type": "application/x-www-form-urlencoded",
                    "Origin": "https://bemanicn.com",
                    "Referer": oauth_response.url,
                    "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                    "sec-ch-ua-mobile": "?0",
                    "sec-ch-ua-platform": "\"Windows\"",
                    "sec-fetch-dest": "document",
                    "sec-fetch-mode": "navigate",
                    "sec-fetch-site": "same-origin",
                    "sec-fetch-user": "?1",
                    "upgrade-insecure-requests": "1"
                }
                
                logger.info("发送授权确认请求...")
                authorize_response = session.post(authorize_url, data=form_data, headers=authorize_headers, allow_redirects=True, timeout=30)
                print_response_info(authorize_response)
                
                # 检查是否包含授权码
                if "u.otogame.net/auth/callback" in authorize_response.url and "code=" in authorize_response.url:
                    logger.info("成功获取授权并重定向到回调URL")
                    callback_url = authorize_response.url
                    parsed_url = urlparse(callback_url)
                    code = parse_qs(parsed_url.query).get('code', [None])[0]
                    
                    if code:
                        logger.info("成功获取授权码")
                        logger.debug(f"授权码: {code[:10]}...")
                        return get_tokens_with_code(session, code)
                    else:
                        logger.error("无法从URL获取授权码")
                        return None
                else:
                    logger.error(f"授权确认后未获取到有效的授权码")
                    logger.error(f"响应URL: {authorize_response.url}")
                    return None

            # 情况2: 如果URL中包含授权码（已授权用户的情况）
            elif "u.otogame.net/auth/callback" in oauth_response.url and "code=" in oauth_response.url:
                logger.info("检测到已授权状态，直接获取授权码")
                callback_url = oauth_response.url
                parsed_url = urlparse(callback_url)
                code = parse_qs(parsed_url.query).get('code', [None])[0]
                
                if code:
                    logger.info("成功获取授权码")
                    logger.debug(f"授权码: {code[:10]}...")
                    return get_tokens_with_code(session, code)
                else:
                    logger.error("无法从URL获取授权码")
                    return None

            # 情况3: 需要登录
            else:
                # 继续原有的登录流程代码
                logger.info("Step 3: 登录到bemanicn.com")
                # 获取XSRF令牌
                cookies = session.cookies.get_dict()
                logger.debug(f"当前Cookie: {cookies}")
                xsrf_token = cookies.get('XSRF-TOKEN')
                
                if not xsrf_token:
                    logger.warning("无法获取XSRF令牌，尝试从响应头或HTML中获取")
                    # 从Set-Cookie头解析
                    if 'Set-Cookie' in oauth_response.headers:
                        logger.debug(f"从Set-Cookie尝试获取XSRF令牌")
                        for cookie in oauth_response.headers.get('Set-Cookie', '').split(';'):
                            if 'XSRF-TOKEN=' in cookie:
                                xsrf_token = cookie.split('XSRF-TOKEN=')[1].split(';')[0]
                                logger.debug(f"从Set-Cookie找到令牌: {xsrf_token}")
                                break
                
                    # 如果还是没有，从HTML中获取
                    if not xsrf_token:
                        logger.debug("从HTML尝试获取CSRF令牌")
                        soup = BeautifulSoup(oauth_response.text, 'html.parser')
                        meta_tag = soup.find('meta', {'name': 'csrf-token'})
                        if meta_tag:
                            xsrf_token = meta_tag['content']
                            logger.debug(f"从HTML找到令牌: {xsrf_token}")
                
                if not xsrf_token:
                    logger.warning("警告：无法获取XSRF令牌，登录可能会失败")
                else:
                    logger.info(f"成功获取XSRF令牌: {xsrf_token[:10]}...")
                
                # 获取X-Inertia-Version (如果页面中有)
                inertia_version = None
                try:
                    soup = BeautifulSoup(oauth_response.text, 'html.parser')
                    scripts = soup.find_all('script')
                    for script in scripts:
                        if script.string and 'Inertia' in script.string and 'version' in script.string:
                            import re
                            match = re.search(r'version:\s*[\'"]([^\'"]+)[\'"]', script.string)
                            if match:
                                inertia_version = match.group(1)
                                logger.debug(f"从页面提取Inertia版本: {inertia_version}")
                                break
                except Exception as e:
                    logger.error(f"解析Inertia版本时出错: {e}")
                
                if not inertia_version:
                    inertia_version = "207fd484b7c2ceeff7800b8c8a11b3b6"  # 使用默认值
                    logger.info(f"使用默认Inertia版本: {inertia_version}")
                else:
                    logger.info(f"成功获取Inertia版本: {inertia_version}")
                
                # 登录请求
                login_url = "https://bemanicn.com/login"
                login_data = {
                    "email": email,
                    "password": password,
                    "remember": "on"
                }
                
                # X-XSRF-TOKEN需要URL解码后的令牌
                import urllib.parse
                decoded_xsrf = urllib.parse.unquote(xsrf_token)
                logger.debug(f"解码后的XSRF令牌: {decoded_xsrf[:10]}...")
                
                login_headers = {
                    "User-Agent": user_agent,
                    "Accept": "text/html, application/xhtml+xml",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Content-Type": "application/json",
                    "Origin": "https://bemanicn.com",
                    "Referer": "https://bemanicn.com/login",
                    "X-Inertia": "true",
                    "X-Inertia-Version": inertia_version,
                    "X-Requested-With": "XMLHttpRequest",
                    "X-XSRF-TOKEN": decoded_xsrf,
                    "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                    "sec-ch-ua-mobile": "?0",
                    "sec-ch-ua-platform": "\"Windows\"",
                    "sec-fetch-dest": "empty",
                    "sec-fetch-mode": "cors",
                    "sec-fetch-site": "same-origin"
                }
                
                logger.debug(f"发送登录请求到: {login_url}")
                logger.debug(f"登录请求头: {login_headers}")
                logger.debug(f"登录数据: {login_data}")
                
                login_response = session.post(login_url, json=login_data, headers=login_headers, timeout=30)
                print_response_info(login_response)
                
                # 检查是否返回了409 Conflict状态，这可能表示需要重定向到OAuth页面
                if login_response.status_code == 409 and 'X-Inertia-Location' in login_response.headers:
                    redirect_to = login_response.headers['X-Inertia-Location']
                    logger.info(f"登录成功，重定向到: {redirect_to}")
                    
                    # 访问重定向URL
                    oauth_redirect_headers = {
                        "User-Agent": user_agent,
                        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                        "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                        "Referer": "https://bemanicn.com/login",
                        "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                        "sec-ch-ua-mobile": "?0",
                        "sec-ch-ua-platform": "\"Windows\"",
                        "sec-fetch-dest": "document",
                        "sec-fetch-mode": "navigate",
                        "sec-fetch-site": "same-origin",
                        "sec-fetch-user": "?1",
                        "upgrade-insecure-requests": "1"
                    }
                    
                    logger.debug(f"跟随重定向到: {redirect_to}")
                    logger.debug(f"当前Cookies: {requests.utils.dict_from_cookiejar(session.cookies)}")
                    
                    # 检查登录后的重定向响应
                    oauth_redirect_response = session.get(redirect_to, headers=oauth_redirect_headers, timeout=30)
                    print_response_info(oauth_redirect_response)
                    
                    # 如果重定向到了回调URL并且包含授权码
                    if "u.otogame.net/auth/callback" in oauth_redirect_response.url and "code=" in oauth_redirect_response.url:
                        logger.info("登录成功并获取到授权码")
                        callback_url = oauth_redirect_response.url
                        parsed_url = urlparse(callback_url)
                        code = parse_qs(parsed_url.query).get('code', [None])[0]
                        
                        if code:
                            logger.info("成功获取授权码")
                            logger.debug(f"授权码: {code[:10]}...")
                            return get_tokens_with_code(session, code)
                        else:
                            logger.error("无法从URL获取授权码")
                            return None
                    # 如果需要确认授权
                    elif "授权提示" in oauth_redirect_response.text and "想要访问您的账户" in oauth_redirect_response.text:
                        logger.info("检测到授权确认页面，正在处理...")
                        
                        # 解析HTML获取表单数据
                        soup = BeautifulSoup(oauth_redirect_response.text, 'html.parser')
                        # 查找同意按钮所在的表单（不含DELETE方法的表单）
                        authorize_form = soup.find('form', {'action': lambda x: x and '/oauth/authorize' in x and not 'DELETE' in str(x)})
                        
                        if not authorize_form:
                            logger.error("无法找到授权表单")
                            return None
                            
                        # 提取表单中的所有隐藏字段
                        form_data = {}
                        for input_tag in authorize_form.find_all('input', {'type': 'hidden'}):
                            if input_tag.get('name') and input_tag.get('value'):
                                form_data[input_tag['name']] = input_tag['value']
                        
                        logger.debug(f"授权表单数据: {form_data}")
                        
                        # 发送同意授权的请求
                        authorize_url = "https://bemanicn.com/oauth/authorize"
                        authorize_headers = {
                            "User-Agent": user_agent,
                            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                            "Content-Type": "application/x-www-form-urlencoded",
                            "Origin": "https://bemanicn.com",
                            "Referer": oauth_redirect_response.url,
                            "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                            "sec-ch-ua-mobile": "?0",
                            "sec-ch-ua-platform": "\"Windows\"",
                            "sec-fetch-dest": "document",
                            "sec-fetch-mode": "navigate",
                            "sec-fetch-site": "same-origin",
                            "sec-fetch-user": "?1",
                            "upgrade-insecure-requests": "1"
                        }
                        
                        logger.info("发送授权确认请求...")
                        authorize_response = session.post(authorize_url, data=form_data, headers=authorize_headers, allow_redirects=True, timeout=30)
                        print_response_info(authorize_response)
                        
                        # 检查是否包含授权码
                        if "u.otogame.net/auth/callback" in authorize_response.url and "code=" in authorize_response.url:
                            logger.info("成功获取授权并重定向到回调URL")
                            callback_url = authorize_response.url
                            parsed_url = urlparse(callback_url)
                            code = parse_qs(parsed_url.query).get('code', [None])[0]
                            
                            if code:
                                logger.info("成功获取授权码")
                                logger.debug(f"授权码: {code[:10]}...")
                                return get_tokens_with_code(session, code)
                            else:
                                logger.error("无法从URL获取授权码")
                                return None
                        else:
                            logger.error(f"授权确认后未获取到有效的授权码")
                            logger.error(f"响应URL: {authorize_response.url}")
                            return None
                    else:
                        logger.error(f"登录后未获取到预期的响应")
                        logger.error(f"响应URL: {oauth_redirect_response.url}")
                        logger.error(f"响应内容: {oauth_redirect_response.text[:200]}...")
                        return None
                
        except json.JSONDecodeError as e:
            logger.error(f"解析JSON响应失败: {e}")
            logger.error(f"原始响应: {response.text[:200]}...")
            return None
            
    except requests.exceptions.RequestException as e:
        logger.error(f"请求过程中出错: {e}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"错误状态码: {e.response.status_code}")
            logger.error(f"错误内容: {e.response.text[:200]}...")
        return None

def get_tokens_with_code(session, code):
    """
    使用授权码获取访问令牌和ID令牌
    """
    logger.info("正在使用授权码获取令牌...")
    user_agent = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36"
    
    try:
        # 首先访问回调URL获取访问令牌
        callback_url = f"https://u.otogame.net/auth/callback?code={code}"
        headers = {
            "User-Agent": user_agent,
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Referer": "https://bemanicn.com/",
            "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-fetch-dest": "document",
            "sec-fetch-mode": "navigate",
            "sec-fetch-site": "cross-site",
            "sec-fetch-user": "?1",
            "upgrade-insecure-requests": "1",
            "priority": "u=0, i"
        }
        
        # 从URL中提取state参数
        parsed_url = urlparse(callback_url)
        state = parse_qs(parsed_url.query).get('state', [None])[0]
        if state:
            # 设置oauthState cookie
            session.cookies.set('oauthState', state.replace('=', ''), domain='u.otogame.net')
        
        logger.debug(f"访问回调URL: {callback_url}")
        logger.debug(f"当前Cookies: {requests.utils.dict_from_cookiejar(session.cookies)}")
        
        response = session.get(callback_url, headers=headers, timeout=30)
        print_response_info(response)
        response.raise_for_status()

        # 调用API获取访问令牌
        api_callback_url = f"https://u.otogame.net/api/aime/user/callback?code={code}&state={state}"
        api_callback_headers = {
            "User-Agent": user_agent,
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
            "Referer": callback_url,
            "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "priority": "u=1, i"
        }
        
        logger.debug(f"调用API获取访问令牌: {api_callback_url}")
        api_callback_response = session.get(api_callback_url, headers=api_callback_headers, timeout=30)
        print_response_info(api_callback_response)
        api_callback_response.raise_for_status()
        
        try:
            token_data = api_callback_response.json()
            if token_data.get("code") != 0:
                logger.error(f"获取访问令牌失败: {token_data}")
                return None
            
            auth_token = token_data.get("data", {}).get("token", {}).get("access_token")
            if not auth_token:
                logger.error("访问令牌为空")
                return None
            
            logger.info("成功获取访问令牌")
            
            # 获取ID令牌
            id_token_url = "https://u.otogame.net/api/aime/token/id"
            id_token_headers = {
                "User-Agent": user_agent,
                "Accept": "application/json, text/plain, */*",
                "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                "Authorization": f"Bearer {auth_token}",
                "Referer": callback_url,
                "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": "\"Windows\"",
                "sec-fetch-dest": "empty",
                "sec-fetch-mode": "cors",
                "sec-fetch-site": "same-origin",
                "priority": "u=1, i"
            }
            
            logger.debug(f"获取ID令牌: {id_token_url}")
            logger.debug(f"使用授权令牌: {auth_token[:50]}...")
            
            id_token_response = session.get(id_token_url, headers=id_token_headers, timeout=30)
            print_response_info(id_token_response)
            id_token_response.raise_for_status()
            
            try:
                id_token_data = id_token_response.json()
                if id_token_data.get("code") != 0:
                    logger.error(f"获取ID令牌失败: {id_token_data}")
                    return None
                    
                id_token = id_token_data.get("data", {}).get("id_token")
                if not id_token:
                    logger.error("ID令牌为空")
                    return None
                    
                logger.info("成功获取ID令牌")
                
                # 访问音乐页面
                music_url = "https://u.otogame.net/ongeki/music"
                music_headers = {
                    "User-Agent": user_agent,
                    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Authorization": f"Bearer {id_token}",
                    "Referer": "https://u.otogame.net/",
                    "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                    "sec-ch-ua-mobile": "?0",
                    "sec-ch-ua-platform": "\"Windows\"",
                    "sec-fetch-dest": "document",
                    "sec-fetch-mode": "navigate",
                    "sec-fetch-site": "same-origin",
                    "sec-fetch-user": "?1",
                    "upgrade-insecure-requests": "1",
                    "priority": "u=0, i"
                }
                
                logger.debug(f"访问音乐页面: {music_url}")
                music_response = session.get(music_url, headers=music_headers, timeout=30)
                print_response_info(music_response)
                
                # 获取评分数据
                rating_url = "https://u.otogame.net/api/game/ongeki/rating"
                rating_headers = {
                    "User-Agent": user_agent,
                    "Accept": "application/json, text/plain, */*",
                    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
                    "Authorization": f"Bearer {id_token}",
                    "Referer": "https://u.otogame.net/ongeki/music",
                    "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
                    "sec-ch-ua-mobile": "?0",
                    "sec-ch-ua-platform": "\"Windows\"",
                    "sec-fetch-dest": "empty",
                    "sec-fetch-mode": "cors",
                    "sec-fetch-site": "same-origin",
                    "priority": "u=1, i"
                }
                
                logger.debug(f"获取评分数据: {rating_url}")
                rating_response = session.get(rating_url, headers=rating_headers, timeout=30)
                print_response_info(rating_response)
                
                if rating_response.status_code == 401:
                    logger.error("获取评分数据失败: 未授权")
                    return None
                    
                rating_response.raise_for_status()
                
                logger.info("成功获取所有令牌和评分数据")
                return {
                    "id_token": id_token,
                    "auth_token": auth_token,
                    "rating_data": rating_response.json()
                }
                
            except json.JSONDecodeError:
                logger.error("解析ID令牌响应失败")
                return None
                
        except json.JSONDecodeError:
            logger.error("解析访问令牌响应失败")
            return None
            
    except requests.exceptions.RequestException as e:
        logger.error(f"获取令牌请求失败: {e}")
        if hasattr(e, 'response') and e.response:
            logger.error(f"错误状态码: {e.response.status_code}")
            logger.error(f"错误内容: {e.response.text[:200]}...")
        return None

def get_rating_data(auth_token):
    """
    使用授权令牌获取评分数据
    """
    logger.info("正在获取评分数据...")
    
    try:
        session = create_session()
        url = "https://u.otogame.net/api/game/ongeki/rating"
        
        headers = {
            "accept": "application/json, text/plain, */*",
            "accept-language": "zh-CN,zh;q=0.9,en;q=0.8",
            "authorization": f"Bearer {auth_token}",
            "priority": "u=1, i",
            "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "Referer": "https://u.otogame.net/ongeki/music",
            "Referrer-Policy": "strict-origin-when-cross-origin"
        }
        
        logger.debug("=== 准备发送评分数据请求 ===")
        logger.debug(f"目标URL: {url}")
        logger.debug(f"使用的授权令牌: {auth_token}")
        logger.debug(f"请求头: {json.dumps(headers, ensure_ascii=False, indent=2)}")
        
        # 发送请求前记录会话状态
        logger.debug(f"当前会话Cookies: {session.cookies.get_dict()}")
        
        # 准备请求对象以便记录详细信息
        req = requests.Request('GET', url, headers=headers)
        prepped = session.prepare_request(req)
        print_request_info(prepped)
        
        # 发送请求
        response = session.send(prepped, timeout=30)
        
        # 记录响应信息
        print_response_info(response)
        
        # 即使状态码不是200，也尝试解析响应
        try:
            response_data = response.json()
            if response.status_code >= 400:
                logger.error("API错误响应:")
                logger.error(f"错误代码: {response_data.get('code')}")
                logger.error(f"错误消息: {response_data.get('message')}")
                logger.error(f"时间戳: {response_data.get('timestamp')}")
            else:
                logger.debug(f"完整响应数据: {json.dumps(response_data, ensure_ascii=False, indent=2)}")
        except Exception as e:
            logger.debug(f"完整响应文本: {response.text}")
            logger.debug(f"解析响应失败: {str(e)}")
        
        response.raise_for_status()
        
        return response.json()
        
    except requests.exceptions.RequestException as e:
        logger.error(f"获取评分数据失败: {e}")
        if hasattr(e, 'response') and e.response:
            try:
                error_data = e.response.json()
                logger.error("API错误详情:")
                logger.error(f"错误代码: {error_data.get('code')}")
                logger.error(f"错误消息: {error_data.get('message')}")
                logger.error(f"时间戳: {error_data.get('timestamp')}")
            except:
                logger.error(f"错误响应文本: {e.response.text}")
            logger.error(f"错误响应头: {json.dumps(dict(e.response.headers), ensure_ascii=False, indent=2)}")
        return None

def get_player_profile(auth_token, session=None):
    """
    使用授权令牌获取玩家资料
    """
    logger.info("正在获取玩家资料...")
    
    try:
        if session is None:
            session = create_session()
        url = "https://u.otogame.net/api/game/ongeki/profile"
        
        headers = {
            "accept": "application/json, text/plain, */*",
            "accept-language": "zh-CN,zh;q=0.9,en;q=0.8",
            "authorization": f"Bearer {auth_token}",
            "priority": "u=1, i",
            "sec-ch-ua": "\"Chromium\";v=\"134\", \"Not:A-Brand\";v=\"24\", \"Google Chrome\";v=\"134\"",
            "sec-ch-ua-mobile": "?0",
            "sec-ch-ua-platform": "\"Windows\"",
            "sec-fetch-dest": "empty",
            "sec-fetch-mode": "cors",
            "sec-fetch-site": "same-origin",
            "Referer": "https://u.otogame.net/ongeki/profile",
            "Referrer-Policy": "strict-origin-when-cross-origin"
        }
        
        logger.debug("=== 准备发送玩家资料请求 ===")
        logger.debug(f"目标URL: {url}")
        logger.debug(f"使用的授权令牌: {auth_token[:20]}...")  # 只显示令牌前20个字符
        logger.debug(f"请求头: {json.dumps(headers, ensure_ascii=False, indent=2)}")
        
        # 发送请求前记录会话状态
        logger.debug(f"当前会话Cookies: {session.cookies.get_dict()}")
        
        # 准备请求对象以便记录详细信息
        req = requests.Request('GET', url, headers=headers)
        prepped = session.prepare_request(req)
        print_request_info(prepped)
        
        # 发送请求
        response = session.send(prepped, timeout=30)
        
        # 记录响应信息
        print_response_info(response)
        
        # 检查响应状态码
        if response.status_code == 401:
            logger.error("获取玩家资料失败：未授权（401）")
            logger.error("这可能是因为令牌已过期或无效")
            return None
        elif response.status_code == 403:
            logger.error("获取玩家资料失败：禁止访问（403）")
            logger.error("这可能是因为没有足够的权限")
            return None
        elif response.status_code != 200:
            logger.error(f"获取玩家资料失败：HTTP {response.status_code}")
            try:
                error_data = response.json()
                logger.error(f"错误详情: {json.dumps(error_data, ensure_ascii=False, indent=2)}")
            except:
                logger.error(f"响应内容: {response.text[:200]}...")
            return None
        
        # 尝试解析响应
        try:
            profile_data = response.json()
            if profile_data.get('code') != "ok":
                logger.error(f"API返回错误: {profile_data.get('message', '未知错误')}")
                return None
                
            # 检查数据结构
            if 'data' not in profile_data:
                logger.error("API响应中缺少data字段")
                logger.debug(f"完整响应: {json.dumps(profile_data, ensure_ascii=False, indent=2)}")
                return None
            
            return profile_data
            
        except json.JSONDecodeError as e:
            logger.error(f"解析响应JSON失败: {e}")
            logger.error(f"响应内容: {response.text[:200]}...")
            return None
            
    except requests.exceptions.RequestException as e:
        logger.error(f"获取玩家资料失败: {e}")
        if hasattr(e, 'response') and e.response:
            try:
                error_data = e.response.json()
                logger.error("API错误详情:")
                logger.error(f"错误代码: {error_data.get('code')}")
                logger.error(f"错误消息: {error_data.get('message')}")
            except:
                logger.error(f"错误响应文本: {e.response.text[:200]}...")
            logger.error(f"错误响应头: {json.dumps(dict(e.response.headers), ensure_ascii=False, indent=2)}")
        return None
    except Exception as e:
        logger.error(f"获取玩家资料时发生未预期的错误: {e}")
        logger.error(traceback.format_exc())
        return None

class B50Converter:
    @staticmethod
    def calculate_constant(score, rating):
        """根据分数和rating计算定数"""
        rating = rating / 100  # 将rating转换为小数形式
        
        if score >= 1007500:
            constant = rating - 2.00
        elif score >= 1000000:
            # 线性内插: 1000000->+1.50, 1007500->+2.00
            position = (score - 1000000) / 7500
            bonus = 1.50 + position * 0.50
            constant = rating - bonus
        elif score >= 990000:
            # 线性内插: 990000->+1.00, 1000000->+1.50
            position = (score - 990000) / 10000
            bonus = 1.00 + position * 0.50
            constant = rating - bonus
        elif score >= 970000:
            # 线性内插: 970000->+0.00, 990000->+1.00
            position = (score - 970000) / 20000
            bonus = position
            constant = rating - bonus
        elif score >= 900000:
            # 线性内插: 900000->-4.00, 970000->+0.00
            position = (score - 900000) / 70000
            bonus = -4.00 + position * 4.00
            constant = rating - bonus
        elif score >= 800000:
            # 线性内插: 800000->-6.00, 900000->-4.00
            position = (score - 800000) / 100000
            bonus = -6.00 + position * 2.00
            constant = rating - bonus
        else:
            constant = rating  # 500000-800000区间为0加成

        # 将定数四舍五入到最近的0.1
        return round(constant * 10) / 10

    @staticmethod
    def get_difficulty_text(diff):
        """获取难度对应的文本"""
        difficulties = {
            0: "BASIC",
            1: "ADVANCE", 
            2: "EXPERT",
            3: "MASTER",
            10: "LUNATIC"
        }
        return difficulties.get(diff, f"未知({diff})")

    @staticmethod
    def set_number_format(cell, value, is_rating=False):
        """设置数字格式，rating显示两位小数，定数显示一位小数"""
        cell.value = value
        if isinstance(value, (int, float)):
            if is_rating:
                cell.number_format = '0.00'  # rating显示两位小数
            elif value == int(value):
                cell.number_format = '0.0'   # 整数也显示一位小数
            else:
                cell.number_format = '0.0'   # 定数显示一位小数

    def convert_to_excel(self, json_file, excel_file):
        """将JSON格式的B50数据转换为Excel格式"""
        logger.info(f"开始将 {json_file} 转换为Excel格式...")
        
        # 读取JSON文件
        with open(json_file, 'r', encoding='utf-8') as f:
            merged_data = json.load(f)
        
        # 获取评分数据
        data = merged_data.get("rating", {})
        profile = merged_data.get("profile", {}).get("data", {})

        # 获取文件中的总rating值
        total_rating = data['data']['rating'] / 100
        best_rating = data['data']['best_rating'] / 100
        new_rating = data['data']['best_new_rating'] / 100
        recent_rating = data['data']['hot_rating'] / 100

        # 处理三个部分的数据
        best_data = []
        for idx, song in enumerate(data['data']['best_rating_list'], 1):
            if song['rating'] > 0:  # 只处理有效的rating
                name = song['music']['name']
                diff_text = self.get_difficulty_text(song['difficulty'])
                score = song['score']
                rating = song['rating'] / 100
                constant = self.calculate_constant(score, song['rating'])
                
                best_data.append({
                    '次序': idx,
                    '曲名': name,
                    '难度': diff_text,
                    '定数': constant,
                    '分数': score,
                    '单曲Rating': rating
                })

        new_data = []
        for idx, song in enumerate(data['data']['best_new_rating_list'], 1):
            if song['rating'] > 0:
                name = song['music']['name']
                diff_text = self.get_difficulty_text(song['difficulty'])
                score = song['score']
                rating = song['rating'] / 100
                constant = self.calculate_constant(score, song['rating'])
                
                new_data.append({
                    '次序': idx,
                    '曲名': name,
                    '难度': diff_text,
                    '定数': constant,
                    '分数': score,
                    '单曲Rating': rating
                })

        recent_data = []
        for idx, song in enumerate(data['data']['hot_rating_list'], 1):
            if song['rating'] > 0:
                name = song['music']['name']
                diff_text = self.get_difficulty_text(song['difficulty'])
                score = song['score']
                rating = song['rating'] / 100
                constant = self.calculate_constant(score, song['rating'])
                
                recent_data.append({
                    '次序': idx,
                    '曲名': name,
                    '难度': diff_text,
                    '定数': constant,
                    '分数': score,
                    '单曲Rating': rating
                })

        # 创建Excel文件
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            
            # 写入B50详情
            writer.sheets['B50详情'] = writer.book.create_sheet('B50详情', 0)
            
            # 写入最佳曲目数据
            row_offset = 1  # 从第2行开始（第1行为标题）
            writer.sheets['B50详情'].merge_cells(f'A{row_offset}:F{row_offset}')
            writer.sheets['B50详情'].cell(row=row_offset, column=1, value="RATING对象曲（最佳）")
            writer.sheets['B50详情'].cell(row=row_offset, column=1).alignment = Alignment(horizontal='center')
            
            # 写入列标题
            headers = ['次序', '曲名', '难度', '定数', '分数', '单曲Rating']
            for col_idx, header in enumerate(headers, 1):
                writer.sheets['B50详情'].cell(row=row_offset+1, column=col_idx, value=header)
            
            # 写入数据
            for idx, row in enumerate(best_data):
                row_idx = row_offset + 2 + idx
                for col_idx, (key, value) in enumerate(row.items(), 1):
                    cell = writer.sheets['B50详情'].cell(row=row_idx, column=col_idx)
                    if key in ['定数', '单曲Rating']:
                        self.set_number_format(cell, value, is_rating=(key=='单曲Rating'))
                    else:
                        cell.value = value
            
            # 写入统计信息
            summary_row = row_offset + 2 + len(best_data)
            writer.sheets['B50详情'].cell(row=summary_row, column=1, value=f"歌曲数: {len(best_data)}首")
            cell = writer.sheets['B50详情'].cell(row=summary_row, column=6)
            self.set_number_format(cell, best_rating, is_rating=True)
            
            # 写入新曲目数据（从最佳下面空一行开始）
            row_offset = summary_row + 1
            writer.sheets['B50详情'].merge_cells(f'A{row_offset}:F{row_offset}')
            writer.sheets['B50详情'].cell(row=row_offset, column=1, value="RATING对象曲（新曲）")
            writer.sheets['B50详情'].cell(row=row_offset, column=1).alignment = Alignment(horizontal='center')
            
            # 写入列标题
            for col_idx, header in enumerate(headers, 1):
                writer.sheets['B50详情'].cell(row=row_offset+1, column=col_idx, value=header)
            
            # 写入数据
            for idx, row in enumerate(new_data):
                row_idx = row_offset + 2 + idx
                for col_idx, (key, value) in enumerate(row.items(), 1):
                    cell = writer.sheets['B50详情'].cell(row=row_idx, column=col_idx)
                    if key in ['定数', '单曲Rating']:
                        self.set_number_format(cell, value, is_rating=(key=='单曲Rating'))
                    else:
                        cell.value = value
            
            # 写入统计信息
            summary_row = row_offset + 2 + len(new_data)
            writer.sheets['B50详情'].cell(row=summary_row, column=1, value=f"歌曲数: {len(new_data)}首")
            cell = writer.sheets['B50详情'].cell(row=summary_row, column=6)
            self.set_number_format(cell, new_rating, is_rating=True)
            
            # 写入最近曲目数据（从新曲下面空一行开始）
            row_offset = summary_row + 1
            writer.sheets['B50详情'].merge_cells(f'A{row_offset}:F{row_offset}')
            writer.sheets['B50详情'].cell(row=row_offset, column=1, value="RATING对象曲（最近）")
            writer.sheets['B50详情'].cell(row=row_offset, column=1).alignment = Alignment(horizontal='center')
            
            # 写入列标题
            for col_idx, header in enumerate(headers, 1):
                writer.sheets['B50详情'].cell(row=row_offset+1, column=col_idx, value=header)
            
            # 写入数据
            for idx, row in enumerate(recent_data):
                row_idx = row_offset + 2 + idx
                for col_idx, (key, value) in enumerate(row.items(), 1):
                    cell = writer.sheets['B50详情'].cell(row=row_idx, column=col_idx)
                    if key in ['定数', '单曲Rating']:
                        self.set_number_format(cell, value, is_rating=(key=='单曲Rating'))
                    else:
                        cell.value = value
            
            # 写入统计信息
            summary_row = row_offset + 2 + len(recent_data)
            writer.sheets['B50详情'].cell(row=summary_row, column=1, value=f"歌曲数: {len(recent_data)}首")
            cell = writer.sheets['B50详情'].cell(row=summary_row, column=6)
            self.set_number_format(cell, recent_rating, is_rating=True)
            
            # 写入总Rating（从最近下面空一行开始）
            row_offset = summary_row + 1
            writer.sheets['B50详情'].cell(row=row_offset, column=1, value="总Rating")
            cell = writer.sheets['B50详情'].cell(row=row_offset, column=6)
            self.set_number_format(cell, total_rating, is_rating=True)

            # 自动调整列宽
            for column in range(1, 7):  # 遍历6列
                max_length = 0
                for row in writer.sheets['B50详情'].rows:
                    cell = row[column-1]
                    if cell.value and not isinstance(cell, MergedCell):
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = max_length + 2
                writer.sheets['B50详情'].column_dimensions[get_column_letter(column)].width = adjusted_width
            # 写入玩家信息
            writer.sheets['玩家信息'] = writer.book.create_sheet('玩家信息', 1)
            if profile:
                info_sheet = writer.sheets['玩家信息']
                info_sheet.cell(row=1, column=1, value="玩家信息")
                info_sheet.merge_cells('A1:B1')
                info_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')
                
                # 写入玩家基本信息
                info_data = [
                    ("玩家名称", profile.get('user_name', 'Unknown')),
                    ("等级", profile.get('level', 'Unknown')),
                    ("游玩次数", profile.get('play_count', 'Unknown')),
                    ("最高Rating", profile.get('highest_rating', 'Unknown')/100 if isinstance(profile.get('highest_rating'), (int, float)) else 'Unknown'),
                    ("当前Rating", profile.get('player_rating', 'Unknown')/100 if isinstance(profile.get('player_rating'), (int, float)) else 'Unknown'),
                    ("总点数", profile.get('total_point', 'Unknown')),
                    ("好友码", profile.get('friend_code', 'Unknown')),
                    ("奖章数", profile.get('medal_count', 'Unknown')),
                    ("战斗点数", profile.get('battle_point', 'Unknown')),
                ]
                
                for idx, (key, value) in enumerate(info_data, 2):
                    info_sheet.cell(row=idx, column=1, value=key)
                    cell = info_sheet.cell(row=idx, column=2)
                    if isinstance(value, float):
                        self.set_number_format(cell, value, is_rating=True)
                    else:
                        cell.value = value
                
                # 调整列宽
                for column in range(1, 3):
                    max_length = 0
                    for row in range(1, len(info_data) + 2):
                        cell = info_sheet.cell(row=row, column=column)
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    adjusted_width = max_length + 2
                    info_sheet.column_dimensions[get_column_letter(column)].width = adjusted_width


        logger.info(f"转换完成！文件已保存为 {excel_file}")
        logger.info(f"玩家总Rating: {total_rating:.2f}")

def main():
    parser = argparse.ArgumentParser(description='获取ONGEKI评分数据')
    parser.add_argument('--email', help='bemanicn.com账号邮箱')
    parser.add_argument('--password', help='bemanicn.com账号密码')
    parser.add_argument('--output', default='b50.json', help='输出JSON文件名')
    parser.add_argument('--excel', action='store_true', help='同时生成Excel文件')
    parser.add_argument('--debug', action='store_true', help='开启详细调试信息')
    parser.add_argument('--no-proxy', action='store_true', help='禁用代理')
    parser.add_argument('--image', action='store_true', help='生成B55图片')
    
    args = parser.parse_args()
    
    # 设置日志级别
    if args.debug:
        logger.setLevel(logging.DEBUG)
    else:
        logger.setLevel(logging.INFO)
        
    # 如果未提供邮箱或密码，交互式获取
    email = args.email
    password = args.password
    
    if not email:
        email = input("请输入bemanicn.com账号邮箱: ")
    if not password:
        import getpass
        password = getpass.getpass("请输入bemanicn.com账号密码(输入的密码不会显示): ")
    
    logger.info("开始获取ONGEKI评分数据...")
    
    # 设置环境变量禁用代理
    if args.no_proxy:
        logger.info("已禁用代理")
        os.environ['HTTP_PROXY'] = ''
        os.environ['HTTPS_PROXY'] = ''
        os.environ['http_proxy'] = ''
        os.environ['https_proxy'] = ''
    
    # 创建session
    session = create_session()
    
    # 登录并获取令牌
    token_data = login_and_get_token(email, password)
    
    if not token_data:
        logger.error("登录失败，无法获取令牌")
        return
    
    # 检查是否直接获取了评分数据
    rating_data = None
    auth_token = None
    id_token = None
    if isinstance(token_data, dict):
        if "rating_data" in token_data:
            rating_data = token_data["rating_data"]
            auth_token = token_data.get("auth_token")
            id_token = token_data.get("id_token")
            logger.info("已从认证过程中获取评分数据")
        elif "auth_token" in token_data:
            auth_token = token_data["auth_token"]
            id_token = token_data.get("id_token")
            # 使用授权令牌获取评分数据
            logger.info("使用授权令牌获取评分数据")
            rating_data = get_rating_data(auth_token)
    else:
        # 假设token_data直接是评分数据
        logger.info("假设返回的数据直接是评分数据")
        rating_data = token_data
    
    if not rating_data:
        logger.error("获取评分数据失败")
        return
    
    # 准备合并的数据
    merged_data = {
        "profile": None,
        "rating": rating_data
    }
    
    # 如果有id_token，获取玩家资料
    if id_token:
        logger.info("获取玩家资料...")
        profile_data = get_player_profile(id_token, session)
        if profile_data:
            logger.info("玩家资料获取成功")
            # 检查profile数据的内容
            try:
                player_name = profile_data.get('data', {}).get('user_name', 'Unknown')
                player_level = profile_data.get('data', {}).get('level', 'Unknown')
                logger.info(f"玩家名称: {player_name}")
                logger.info(f"玩家等级: {player_level}")
                # 将玩家资料添加到合并数据中
                merged_data["profile"] = profile_data
            except Exception as e:
                logger.warning(f"解析玩家资料时出错: {e}")
        else:
            logger.warning("获取玩家资料失败")
    
    # 保存合并后的数据到文件
    try:
        with open(args.output, 'w', encoding='utf-8') as f:
            json.dump(merged_data, f, ensure_ascii=False, indent=2)
        logger.info(f"数据已保存到 {args.output}")
        save_success = True
    except Exception as e:
        logger.error(f"保存数据失败: {e}")
        save_success = False
    
    if save_success and args.excel:
        try:
            excel_file = args.output.replace('.json', '.xlsx')
            logger.info(f"转换为Excel文件: {excel_file}")
            converter = B50Converter()
            converter.convert_to_excel(args.output, excel_file)
            logger.info(f"Excel文件已生成: {excel_file}")
        except Exception as e:
            logger.error(f"转换为Excel失败: {e}")
    
    if save_success and args.image:
        try:
            logger.info("开始生成B55图片...")
            from b55_gram import B55GramGenerator
            generator = B55GramGenerator()
            image = generator.generate(rating_data, merged_data.get("profile"))
            image.save('b55_gram.png')
            logger.info("B55图片已生成: b55_gram.png")
        except Exception as e:
            logger.error(f"生成B55图片失败: {e}")
            logger.error(traceback.format_exc())
    
    logger.info("操作完成")

if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        logger.info("程序被用户中断")
    except Exception as e:
        logger.error(f"程序出错: {e}")
        logger.error(traceback.format_exc()) 