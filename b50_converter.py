import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# 根据分数和rating计算定数
def calculate_constant(score, rating):
    rating = rating / 100  # 将rating转换为小数形式
    
    if score >= 1007500:
        constant = rating - 2.00
    elif score >= 1000000:
        # 线性内插: 1000000->+1.50, 1007500->+2.00
        position = (score - 1000000) / 7500
        bonus = 1.50 + position * 0.50
        constant = rating - bonus
    elif score >= 990000:
        # 线性内插: 990000->+1.00, 1000000->+1.50
        position = (score - 990000) / 10000
        bonus = 1.00 + position * 0.50
        constant = rating - bonus
    elif score >= 970000:
        # 线性内插: 970000->+0.00, 990000->+1.00
        position = (score - 970000) / 20000
        bonus = position
        constant = rating - bonus
    elif score >= 900000:
        # 线性内插: 900000->-4.00, 970000->+0.00
        position = (score - 900000) / 70000
        bonus = -4.00 + position * 4.00
        constant = rating - bonus
    elif score >= 800000:
        # 线性内插: 800000->-6.00, 900000->-4.00
        position = (score - 800000) / 100000
        bonus = -6.00 + position * 2.00
        constant = rating - bonus
    else:
        constant = rating  # 500000-800000区间为0加成

    # 将定数四舍五入到最近的0.1
    return round(constant * 10) / 10

def set_number_format(cell, value, is_rating=False):
    """设置数字格式，rating显示两位小数，定数显示一位小数"""
    cell.value = value
    if isinstance(value, (int, float)):
        if is_rating:
            cell.number_format = '0.00'  # rating显示两位小数
        elif value == int(value):
            cell.number_format = '0.0'   # 整数也显示一位小数
        else:
            cell.number_format = '0.0'   # 定数显示一位小数

# 获取难度对应的文本
def get_difficulty_text(diff):
    difficulties = {
        0: "BASIC",
        1: "ADVANCE", 
        2: "EXPERT",
        3: "MASTER",
        10: "LUNATIC"
    }
    return difficulties.get(diff, f"未知({diff})")

# 读取JSON文件
with open('b50.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

# 获取文件中的总rating值
total_rating = data['data']['rating'] / 100
best_rating = data['data']['best_rating'] / 100
new_rating = data['data']['best_new_rating'] / 100
recent_rating = data['data']['hot_rating'] / 100

# 处理三个部分的数据
best_data = []
for idx, song in enumerate(data['data']['best_rating_list'], 1):
    if song['rating'] > 0:  # 只处理有效的rating
        name = song['music']['name']
        diff_text = get_difficulty_text(song['difficulty'])
        score = song['score']
        rating = song['rating'] / 100
        constant = calculate_constant(score, song['rating'])
        
        best_data.append({
            '次序': idx,
            '曲名': name,
            '难度': diff_text,
            '定数': constant,
            '分数': score,
            '单曲Rating': rating
        })

new_data = []
for idx, song in enumerate(data['data']['best_new_rating_list'], 1):
    if song['rating'] > 0:
        name = song['music']['name']
        diff_text = get_difficulty_text(song['difficulty'])
        score = song['score']
        rating = song['rating'] / 100
        constant = calculate_constant(score, song['rating'])
        
        new_data.append({
            '次序': idx,
            '曲名': name,
            '难度': diff_text,
            '定数': constant,
            '分数': score,
            '单曲Rating': rating
        })

recent_data = []
for idx, song in enumerate(data['data']['hot_rating_list'], 1):
    if song['rating'] > 0:
        name = song['music']['name']
        diff_text = get_difficulty_text(song['difficulty'])
        score = song['score']
        rating = song['rating'] / 100
        constant = calculate_constant(score, song['rating'])
        
        recent_data.append({
            '次序': idx,
            '曲名': name,
            '难度': diff_text,
            '定数': constant,
            '分数': score,
            '单曲Rating': rating
        })

# 创建Excel文件
with pd.ExcelWriter('b50.xlsx', engine='openpyxl') as writer:
    # 写入标题行
    writer.sheets['B50详情'] = writer.book.create_sheet('B50详情')
    
    # 写入最佳曲目数据
    row_offset = 1  # 从第2行开始（第1行为标题）
    writer.sheets['B50详情'].merge_cells(f'A{row_offset}:F{row_offset}')
    writer.sheets['B50详情'].cell(row=row_offset, column=1, value="RATING对象曲（最佳）")
    writer.sheets['B50详情'].cell(row=row_offset, column=1).alignment = Alignment(horizontal='center')
    
    # 写入列标题
    headers = ['次序', '曲名', '难度', '定数', '分数', '单曲Rating']
    for col_idx, header in enumerate(headers, 1):
        writer.sheets['B50详情'].cell(row=row_offset+1, column=col_idx, value=header)
    
    # 写入数据
    for idx, row in enumerate(best_data):
        row_idx = row_offset + 2 + idx
        for col_idx, (key, value) in enumerate(row.items(), 1):
            cell = writer.sheets['B50详情'].cell(row=row_idx, column=col_idx)
            if key in ['定数', '单曲Rating']:
                set_number_format(cell, value, is_rating=(key=='单曲Rating'))
            else:
                cell.value = value
    
    # 写入统计信息
    summary_row = row_offset + 2 + len(best_data)
    writer.sheets['B50详情'].cell(row=summary_row, column=1, value=f"歌曲数: {len(best_data)}首")
    cell = writer.sheets['B50详情'].cell(row=summary_row, column=6)
    set_number_format(cell, best_rating, is_rating=True)
    
    # 写入新曲目数据（从最佳下面空一行开始）
    row_offset = summary_row + 1
    writer.sheets['B50详情'].merge_cells(f'A{row_offset}:F{row_offset}')
    writer.sheets['B50详情'].cell(row=row_offset, column=1, value="RATING对象曲（新曲）")
    writer.sheets['B50详情'].cell(row=row_offset, column=1).alignment = Alignment(horizontal='center')
    
    # 写入列标题
    for col_idx, header in enumerate(headers, 1):
        writer.sheets['B50详情'].cell(row=row_offset+1, column=col_idx, value=header)
    
    # 写入数据
    for idx, row in enumerate(new_data):
        row_idx = row_offset + 2 + idx
        for col_idx, (key, value) in enumerate(row.items(), 1):
            cell = writer.sheets['B50详情'].cell(row=row_idx, column=col_idx)
            if key in ['定数', '单曲Rating']:
                set_number_format(cell, value, is_rating=(key=='单曲Rating'))
            else:
                cell.value = value
    
    # 写入统计信息
    summary_row = row_offset + 2 + len(new_data)
    writer.sheets['B50详情'].cell(row=summary_row, column=1, value=f"歌曲数: {len(new_data)}首")
    cell = writer.sheets['B50详情'].cell(row=summary_row, column=6)
    set_number_format(cell, new_rating, is_rating=True)
    
    # 写入最近曲目数据（从新曲下面空一行开始）
    row_offset = summary_row + 1
    writer.sheets['B50详情'].merge_cells(f'A{row_offset}:F{row_offset}')
    writer.sheets['B50详情'].cell(row=row_offset, column=1, value="RATING对象曲（最近）")
    writer.sheets['B50详情'].cell(row=row_offset, column=1).alignment = Alignment(horizontal='center')
    
    # 写入列标题
    for col_idx, header in enumerate(headers, 1):
        writer.sheets['B50详情'].cell(row=row_offset+1, column=col_idx, value=header)
    
    # 写入数据
    for idx, row in enumerate(recent_data):
        row_idx = row_offset + 2 + idx
        for col_idx, (key, value) in enumerate(row.items(), 1):
            cell = writer.sheets['B50详情'].cell(row=row_idx, column=col_idx)
            if key in ['定数', '单曲Rating']:
                set_number_format(cell, value, is_rating=(key=='单曲Rating'))
            else:
                cell.value = value
    
    # 写入统计信息
    summary_row = row_offset + 2 + len(recent_data)
    writer.sheets['B50详情'].cell(row=summary_row, column=1, value=f"歌曲数: {len(recent_data)}首")
    cell = writer.sheets['B50详情'].cell(row=summary_row, column=6)
    set_number_format(cell, recent_rating, is_rating=True)
    
    # 写入总Rating（从最近下面空一行开始）
    row_offset = summary_row + 1
    writer.sheets['B50详情'].cell(row=row_offset, column=1, value="总Rating")
    cell = writer.sheets['B50详情'].cell(row=row_offset, column=6)
    set_number_format(cell, total_rating, is_rating=True)

    # 自动调整列宽
    for column in range(1, 7):  # 遍历6列
        max_length = 0
        for row in writer.sheets['B50详情'].rows:
            cell = row[column-1]
            if cell.value and not isinstance(cell, MergedCell):
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        writer.sheets['B50详情'].column_dimensions[get_column_letter(column)].width = adjusted_width

print(f"转换完成！文件已保存为 b50.xlsx")
print(f"玩家总Rating: {total_rating:.2f}") 